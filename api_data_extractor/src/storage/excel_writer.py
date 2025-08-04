# excel_writer.py

"""
This module is responsible for saving structured data into an Excel file (XLSX format).
It includes functions that take data in a specified format and write it to an Excel file,
ensuring proper formatting and organization.
"""

import openpyxl
from openpyxl.utils import get_column_letter

def save_to_excel(data, file_path='data/output.xlsx'):
    """
    Saves the provided data to an Excel file.

    Parameters:
    data (list of dict): The data to be saved, where each dictionary represents a row.
    file_path (str): The path to the output Excel file. Default is 'data/output.xlsx'.

    Returns:
    None
    """
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write the header row
    if data:
        headers = data[0].keys()
        worksheet.append(headers)

        # Write the data rows
        for item in data:
            worksheet.append(item.values())

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Save the workbook to the specified file path
    workbook.save(file_path)